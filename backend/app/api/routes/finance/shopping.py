from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.api.crud import crud_router
from app.core.database import get_db
from app.core.security import get_current_user
from app.models import (
    FinanceShoppingLedger,
    FinanceShoppingPlatform,
    FinanceShoppingRecord,
    UserProfile,
)
from app.schemas.finance import (
    ShoppingCreate,
    ShoppingLedgerCreate,
    ShoppingLedgerRead,
    ShoppingPlatformCreate,
    ShoppingPlatformRead,
    ShoppingRead,
)
from app.schemas.health import PageOut

# 平台 / 账本：标准 CRUD
platforms_router = crud_router(
    prefix="/finance/shopping/platforms",
    tag="finance-shopping-platforms",
    model=FinanceShoppingPlatform,
    create_schema=ShoppingPlatformCreate,
    read_schema=ShoppingPlatformRead,
    order_by=FinanceShoppingPlatform.id,
)

ledgers_router = crud_router(
    prefix="/finance/shopping/ledgers",
    tag="finance-shopping-ledgers",
    model=FinanceShoppingLedger,
    create_schema=ShoppingLedgerCreate,
    read_schema=ShoppingLedgerRead,
    order_by=FinanceShoppingLedger.id,
)


# 记录：自定义路由（支持账本过滤 / 统计 / 导入）
records_router = APIRouter(prefix="/finance/shopping/records", tags=["finance-shopping"])


@records_router.get("", response_model=PageOut[ShoppingRead])
def list_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    start: date | None = None,
    end: date | None = None,
    ledger_id: int | None = None,
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
):
    stmt = select(FinanceShoppingRecord).where(FinanceShoppingRecord.user_id == user.id)
    if start:
        stmt = stmt.where(FinanceShoppingRecord.record_date >= start)
    if end:
        stmt = stmt.where(FinanceShoppingRecord.record_date <= end)
    if ledger_id:
        stmt = stmt.where(FinanceShoppingRecord.ledger_id == ledger_id)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(
        stmt.order_by(FinanceShoppingRecord.record_date.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return PageOut(items=rows, total=total, page=page, page_size=page_size)


@records_router.get("/stats")
def records_stats(
    days: int = Query(90, ge=1, le=3650),
    ledger_id: int | None = None,
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
) -> dict:
    """购物统计：消费总额、按平台、按账本、月度趋势。"""
    since = date.today() - timedelta(days=days - 1)
    stmt = select(FinanceShoppingRecord).where(
        FinanceShoppingRecord.user_id == user.id,
        FinanceShoppingRecord.record_date >= since,
    )
    if ledger_id:
        stmt = stmt.where(FinanceShoppingRecord.ledger_id == ledger_id)
    rows = db.scalars(stmt).all()

    total = sum(r.total_price for r in rows)
    monthly: dict[str, float] = defaultdict(float)
    by_platform: dict[int, float] = defaultdict(float)
    by_ledger: dict[int, float] = defaultdict(float)

    platform_names = {
        p.id: p.name
        for p in db.scalars(
            select(FinanceShoppingPlatform).where(FinanceShoppingPlatform.user_id == user.id)
        ).all()
    }
    ledger_names = {
        l.id: l.name
        for l in db.scalars(
            select(FinanceShoppingLedger).where(FinanceShoppingLedger.user_id == user.id)
        ).all()
    }

    for r in rows:
        monthly[r.record_date.strftime("%Y-%m")] += r.total_price
        if r.platform_id:
            by_platform[r.platform_id] += r.total_price
        if r.ledger_id:
            by_ledger[r.ledger_id] += r.total_price

    return {
        "total": round(total, 2),
        "count": len(rows),
        "monthly_trend": [
            {"month": m, "amount": round(a, 2)} for m, a in sorted(monthly.items())
        ],
        "by_platform": [
            {"platform_id": pid, "platform": platform_names.get(pid, "未分类"), "amount": round(a, 2)}
            for pid, a in sorted(by_platform.items(), key=lambda x: -x[1])
        ],
        "by_ledger": [
            {"ledger_id": lid, "ledger": ledger_names.get(lid, "未分类"), "amount": round(a, 2)}
            for lid, a in sorted(by_ledger.items(), key=lambda x: -x[1])
        ],
    }


@records_router.post("", response_model=ShoppingRead, status_code=201)
def create_record(
    payload: ShoppingCreate,
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
):
    obj = FinanceShoppingRecord(**payload.model_dump(), user_id=user.id)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@records_router.put("/{item_id}", response_model=ShoppingRead)
def update_record(
    item_id: int,
    payload: ShoppingCreate,
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
):
    obj = db.get(FinanceShoppingRecord, item_id)
    if not obj or obj.user_id != user.id:
        raise HTTPException(status_code=404, detail="记录不存在")
    for key, value in payload.model_dump().items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj


@records_router.delete("/{item_id}", status_code=204)
def delete_record(
    item_id: int,
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
):
    obj = db.get(FinanceShoppingRecord, item_id)
    if not obj or obj.user_id != user.id:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(obj)
    db.commit()
    return None


import_router = APIRouter(prefix="/finance/shopping", tags=["finance-shopping"])

# 表头别名映射：兼容常见 xlsx 表头（如"名称/价格/购买平台"等）
_FIELD_ALIASES = {
    "record_date": ["日期", "购买日期"],
    "platform_name": ["平台", "购买平台", "商城"],
    "product_name": ["商品名称", "名称", "商品"],
    "spec": ["规格"],
    "total_price": ["总价", "价格", "金额"],
    "unit_price": ["单价"],
    "order_no": ["订单号"],
    "ledger_name": ["账本"],
}


def _resolve_name(name: str, cache: dict[str, int], model, attr: str, db: Session, user_id: int) -> int | None:
    """按名称解析 id，未收录则自动创建记录。"""
    if not name:
        return None
    if name not in cache:
        obj = model(**{attr: name}, user_id=user_id)
        db.add(obj)
        db.flush()
        cache[name] = obj.id
    return cache[name]


def _coerce_date(value) -> date:
    """尝试把 Excel 日期（datetime / date / 数字 / 分隔字符串）转成 date。"""
    import re

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return date.today()
    # 纯数字：yyyymmdd(8位) 或 yymmdd(6位，如 260902)
    if s.isdigit():
        try:
            if len(s) >= 8:
                return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
            if len(s) == 6:
                return date(2000 + int(s[:2]), int(s[2:4]), int(s[4:6]))
        except ValueError:
            pass
    # 分隔形式：2026-9-3 / 2026/9/3 / 2026年9月3日
    m = re.match(r"^(\d{4})[-/.年月](\d{1,2})[-/.月日]", s)
    if m:
        try:
            return date(*(int(x) for x in m.groups()))
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        try:
            return date.fromisoformat(s)
        except ValueError:
            return date.today()


@import_router.post("/import")
async def import_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
):
    """导入 xlsx 购物记录。表头：日期/平台/商品名称/规格/总价/单价/订单号/账本。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl 依赖")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件：{exc}")

    platform_cache = {
        p.name: p.id
        for p in db.scalars(
            select(FinanceShoppingPlatform).where(FinanceShoppingPlatform.user_id == user.id)
        ).all()
    }
    ledger_cache = {
        l.name: l.id
        for l in db.scalars(
            select(FinanceShoppingLedger).where(FinanceShoppingLedger.user_id == user.id)
        ).all()
    }

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0}

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if not h:
            continue
        for field, aliases in _FIELD_ALIASES.items():
            if h in aliases and field not in col_idx:
                col_idx[field] = i

    def get(raw, key):
        i = col_idx.get(key)
        if i is None or i >= len(raw):
            return None
        return raw[i]

    records = []
    skipped = 0
    for raw in rows[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        rec_date = get(raw, "record_date")
        product = get(raw, "product_name")
        total_cost = get(raw, "total_price")
        if not rec_date or not product or total_cost is None:
            skipped += 1
            continue

        platform_name = str(get(raw, "platform_name")).strip() if get(raw, "platform_name") else ""
        ledger_name = str(get(raw, "ledger_name")).strip() if get(raw, "ledger_name") else ""
        spec = str(get(raw, "spec")).strip() if get(raw, "spec") else None
        order_no = str(get(raw, "order_no")).strip() if get(raw, "order_no") else None
        unit = get(raw, "unit_price")

        records.append(
            FinanceShoppingRecord(
                record_date=_coerce_date(rec_date),
                platform_id=_resolve_name(platform_name, platform_cache, FinanceShoppingPlatform, "name", db, user.id),
                product_name=str(product).strip(),
                spec=spec,
                total_price=round(float(total_cost), 2),
                unit_price=round(float(unit), 2) if unit is not None else None,
                order_no=order_no,
                ledger_id=_resolve_name(ledger_name, ledger_cache, FinanceShoppingLedger, "name", db, user.id),
                note=None,
                user_id=user.id,
            )
        )

    if records:
        db.add_all(records)
        db.commit()
    return {"imported": len(records), "skipped": skipped}