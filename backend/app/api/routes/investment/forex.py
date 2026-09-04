from collections import defaultdict
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models import InvestmentForex, InvestmentFundRecord, UserProfile
from app.schemas.investment import ForexCreate, ForexRead
from app.schemas.health import PageOut

router = APIRouter(prefix="/investment/forex", tags=["investment-forex"])


# --------------------------------------------------------------------------
# 解析/计算辅助
# --------------------------------------------------------------------------
def _coerce_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return date.today()


def _parse_dt(value) -> datetime | None:
    """把 Excel/Mt5 时间细胞解析成 datetime。"""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time(0, 0))
    s = str(value).strip()
    for fmt in (
        "%Y.%m.%d %H:%M:%S",
        "%Y.%m.%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


def _num(value, default: float | None = None):
    if value is None:
        return default
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def _holding_minutes(open_dt: datetime | None, close_dt: datetime | None) -> int | None:
    if not open_dt or not close_dt or close_dt <= open_dt:
        return None
    return int(round((close_dt - open_dt).total_seconds() / 60))


def _parse_time_only(value) -> time | None:
    """把纯时间字符串 %H:%M:%S / %H:%M 解析成 time。"""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    s = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _parse_duration_minutes(value) -> int | None:
    """把持仓时长 H:M:S / M:S 解析成分钟。"""
    if value is None or str(value).strip() == "":
        return None
    s = str(value).strip()
    parts = s.split(":")
    try:
        if len(parts) == 3:
            return int(parts[0]) * 60 + int(parts[1]) + round(int(parts[2]) / 60)
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, TypeError):
        pass
    # 可能是浮点天数（Excel）或分钟数
    try:
        f = float(s)
    except ValueError:
        return None
    if 0 < abs(f) < 1:
        return int(round(f * 24 * 60))
    return int(round(f))


def _combine_open_close(day: date, open_t: str, close_t: str) -> tuple[datetime | None, datetime | None]:
    ot = _parse_time_only(open_t)
    ct = _parse_time_only(close_t)
    if ot is None:
        return None, None
    open_dt = datetime.combine(day, ot)
    if ct is None:
        return open_dt, None
    close_dt = datetime.combine(day, ct)
    if close_dt < open_dt:  # 跨日平仓
        close_dt += timedelta(days=1)
    return open_dt, close_dt


def _trade_net(t: InvestmentForex) -> float:
    """单笔净盈亏 = 盈亏 + 手续费 + 隔夜费（带符号）。"""
    return (t.pnl or 0) + (t.commission or 0) + (t.overnight_fee or 0)


def _trade_date(t: InvestmentForex) -> date:
    if t.close_time:
        return t.close_time.date()
    if t.open_time:
        return t.open_time.date()
    return t.trade_date


# --------------------------------------------------------------------------
# CRUD
# --------------------------------------------------------------------------
def _read(t: InvestmentForex) -> ForexRead:
    return ForexRead.model_validate(t)


@router.get("", response_model=PageOut[ForexRead])
def list_items(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    start: date | None = None,
    end: date | None = None,
    symbol: str | None = None,
    order_type: str | None = None,
    db: Session = Depends(get_db),
    current_user: UserProfile = Depends(get_current_user),
):
    stmt = select(InvestmentForex).where(InvestmentForex.user_id == current_user.id)
    if start:
        stmt = stmt.where(InvestmentForex.trade_date >= start)
    if end:
        stmt = stmt.where(InvestmentForex.trade_date <= end)
    if symbol:
        stmt = stmt.where(InvestmentForex.symbol == symbol)
    if order_type:
        stmt = stmt.where(InvestmentForex.order_type == order_type)
    total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    rows = db.scalars(
        stmt.order_by(InvestmentForex.trade_date.desc(), InvestmentForex.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return PageOut(items=[_read(r) for r in rows], total=total, page=page, page_size=page_size)


def _apply_payload(obj: InvestmentForex, payload: ForexCreate):
    data = payload.model_dump()
    data.setdefault("holding", _holding_minutes(data.get("open_time"), data.get("close_time")))
    # 保底：持仓时间由开平仓时间自动计算
    data["holding"] = data.get("holding") or _holding_minutes(data.get("open_time"), data.get("close_time"))
    for key, value in data.items():
        setattr(obj, key, value)


@router.post("", response_model=ForexRead, status_code=201)
def create_item(payload: ForexCreate, db: Session = Depends(get_db),
                current_user: UserProfile = Depends(get_current_user)):
    obj = InvestmentForex(user_id=current_user.id)
    _apply_payload(obj, payload)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return _read(obj)


@router.put("/{item_id}", response_model=ForexRead)
def update_item(item_id: int, payload: ForexCreate, db: Session = Depends(get_db),
                current_user: UserProfile = Depends(get_current_user)):
    obj = db.scalar(
        select(InvestmentForex).where(
            InvestmentForex.id == item_id,
            InvestmentForex.user_id == current_user.id,
        )
    )
    if not obj:
        raise HTTPException(status_code=404, detail="记录不存在")
    _apply_payload(obj, payload)
    db.commit()
    db.refresh(obj)
    return _read(obj)


@router.delete("/{item_id}", status_code=204)
def delete_item(item_id: int, db: Session = Depends(get_db),
                current_user: UserProfile = Depends(get_current_user)):
    obj = db.scalar(
        select(InvestmentForex).where(
            InvestmentForex.id == item_id,
            InvestmentForex.user_id == current_user.id,
        )
    )
    if not obj:
        raise HTTPException(status_code=404, detail="记录不存在")
    db.delete(obj)
    db.commit()
    return None


# --------------------------------------------------------------------------
# 统计 / 数据分析
# --------------------------------------------------------------------------
def compute_forex_stats(db: Session, days: int = 365, user_id: int | None = None) -> dict:
    stat_trades = select(InvestmentForex)
    stat_funds = select(InvestmentFundRecord)
    if user_id is not None:
        stat_trades = stat_trades.where(InvestmentForex.user_id == user_id)
        stat_funds = stat_funds.where(InvestmentFundRecord.user_id == user_id)
    trades = db.scalars(stat_trades).all()
    closed = [t for t in trades if t.status == "closed"]
    funds = db.scalars(stat_funds).all()
    deposits = sum(f.amount for f in funds if f.record_type == "deposit")
    withdrawals = sum(f.amount for f in funds if f.record_type == "withdraw")

    # 体验金拆三类（遵循 bonus 语义）：
    #   赠金发放 = experience 正数；bonus_loss / bonus_expired = experience 负数（按 remark 前缀区分）
    #   net_capital = deposit + bonus - bonus_loss - withdrawal（bonus_expired 不进净值）
    #   remaining_bonus = max(0, bonus - bonus_loss - bonus_expired)  仅做体验金余额截断
    bonus_credit = 0.0
    bonus_loss = 0.0
    bonus_expired = 0.0
    for f in funds:
        if f.record_type != "experience":
            continue
        if f.amount >= 0:
            bonus_credit += f.amount
        elif f.note and "bns807" in f.note:  # 体验金失效
            bonus_expired += -f.amount
        else:  # 体验金亏损
            bonus_loss += -f.amount
    net_capital = deposits - withdrawals + bonus_credit - bonus_loss
    remaining_bonus = max(0.0, bonus_credit - bonus_loss - bonus_expired)

    nets = {t.id: _trade_net(t) for t in closed}
    gross_pnl = sum(t.pnl or 0 for t in closed)
    total_commission = round(sum(t.commission or 0 for t in closed), 2)
    total_overnight = round(sum(t.overnight_fee or 0 for t in closed), 2)
    net_profit = round(sum(nets.values()), 2)
    account_value = round(net_capital + remaining_bonus + net_profit, 2)

    wins = [n for n in nets.values() if n > 0]
    losses = [n for n in nets.values() if n < 0]
    gross_win = sum(wins)
    gross_loss = abs(sum(losses))
    win_rate = round(len(wins) / len(closed) * 100, 1) if closed else None
    profit_loss_ratio = round(gross_win / gross_loss, 2) if gross_loss else None
    profit_factor = round(gross_win / gross_loss, 2) if gross_loss else None
    avg_win = round(gross_win / len(wins), 2) if wins else None
    avg_loss = round(gross_loss / len(losses), 2) if losses else None

    # 收益曲线（累计净盈亏，按交易日）
    daily_net: dict[date, float] = defaultdict(float)
    for t in closed:
        daily_net[_trade_date(t)] += nets[t.id]
    cumulative, curve = 0.0, []
    for d in sorted(daily_net):
        cumulative += daily_net[d]
        cumulative = round(cumulative, 2)
        # 只保留近 days 天窗口
        if d >= date.today() - timedelta(days=days - 1):
            curve.append(
                {"date": d.isoformat(), "pnl": cumulative,
                 "pos": round(max(0, cumulative), 2), "neg": round(min(0, cumulative), 2)}
            )

    # 按交易品种
    by_symbol: dict[str, dict] = {}
    for t in closed:
        s = by_symbol.setdefault(t.symbol, {"count": 0, "win": 0, "pnl": 0.0})
        s["count"] += 1
        s["pnl"] = round(s["pnl"] + nets[t.id], 2)
        if nets[t.id] > 0:
            s["win"] += 1
    by_symbol_list = [
        {"symbol": k, "count": v["count"],
         "win_rate": round(v["win"] / v["count"] * 100, 1) if v["count"] else 0,
         "pnl": v["pnl"]}
        for k, v in sorted(by_symbol.items(), key=lambda x: -x[1]["pnl"])
    ]
    active_symbols = sorted({t.symbol for t in trades})

    # 数据分析：最大回撤、连胜/连败、平均持仓、时段分布
    drawdown_peak, drawdown_max, dd_pct = 0.0, 0.0, 0.0
    run_total = 0.0
    for t in closed:
        run_total += nets[t.id]
        if run_total > drawdown_peak:
            drawdown_peak = run_total
        if drawdown_peak > 0:
            dd_pct_i = (drawdown_peak - run_total) / drawdown_peak * 100
            if dd_pct_i > dd_pct:
                dd_pct = dd_pct_i
        dd = drawdown_peak - run_total
        if dd > drawdown_max:
            drawdown_max = dd

    longest_win = longest_loss = cur_streak = 0
    cur_len = 0
    for t in closed:
        if nets[t.id] > 0:
            if cur_streak and cur_streak < 0:
                cur_len = 0
            cur_streak = 1
            cur_len += 1
            longest_win = max(longest_win, cur_len)
        else:
            if cur_streak and cur_streak > 0:
                cur_len = 0
            cur_streak = -1
            cur_len += 1
            longest_loss = max(longest_loss, cur_len)

    holding_list = [_holding_minutes(t.open_time, t.close_time) or t.holding or 0 for t in closed if t.close_time]
    avg_holding = round(sum(holding_list) / len(holding_list)) if holding_list else None

    hour_dist: dict[int, int] = defaultdict(int)
    for t in closed:
        pt = t.open_time or t.close_time
        if pt:
            hour_dist[pt.hour] += 1

    # 交易品种盈亏按日、净收益、手续费、隔夜费
    trend_window = [t for t in closed if _trade_date(t) >= date.today() - timedelta(days=days - 1)]
    daily_pnl: dict[date, float] = defaultdict(float)
    for t in trend_window:
        daily_pnl[_trade_date(t)] = round(daily_pnl[_trade_date(t)] + nets[t.id], 2)

    return {
        "summary": {
            "account_value": account_value,
            "net_profit": net_profit,
            "gross_pnl": round(gross_pnl, 2),
            "total_commission": total_commission,
            "total_overnight": total_overnight,
            "trade_count": len(closed),
            "open_count": sum(1 for t in trades if t.status == "open"),
            "win_rate": win_rate,
            "profit_loss_ratio": profit_loss_ratio,
            "profit_factor": profit_factor,
            "total_deposit": round(deposits, 2),
            "total_withdraw": round(withdrawals, 2),
            "total_experience": round(remaining_bonus, 2),
            "symbol_count": len(active_symbols),
        },
        "equity_trend": curve,
        "daily_pnl": [{"date": d.isoformat(), "amount": v} for d, v in sorted(daily_pnl.items())],
        "by_symbol": by_symbol_list,
        "symbols": active_symbols,
        "analysis": {
            "avg_win": avg_win,
            "avg_loss": avg_loss,
            "max_drawdown": round(drawdown_max, 2),
            "max_drawdown_pct": round(dd_pct, 2) if drawdown_peak else 0,
            "profit_factor": profit_factor,
            "longest_win_streak": longest_win,
            "longest_loss_streak": longest_loss,
            "avg_holding_minutes": avg_holding,
            "hour_dist": [
                {"hour": h, "count": c} for h, c in sorted(hour_dist.items())
            ],
        },
    }


@router.get("/stats")
def stats(days: int = Query(365, ge=1, le=3650), db: Session = Depends(get_db),
          current_user: UserProfile = Depends(get_current_user)):
    return compute_forex_stats(db, days, current_user.id)


# --------------------------------------------------------------------------
# 交易日历（按月：哪些天有交易、盈亏、笔数）
# --------------------------------------------------------------------------
@router.get("/calendar")
def calendar(month: str | None = Query(None), db: Session = Depends(get_db),
             current_user: UserProfile = Depends(get_current_user)):
    today = date.today()
    if month:
        try:
            y, m = month.split("-")
            y, m = int(y), int(m)
        except (ValueError, AttributeError):
            y, m = today.year, today.month
    else:
        y, m = today.year, today.month
    import calendar as _cal

    start = date(y, m, 1)
    end = date(y, m, _cal.monthrange(y, m)[1])
    trades = db.scalars(
        select(InvestmentForex).where(
            InvestmentForex.user_id == current_user.id,
            InvestmentForex.trade_date >= start,
            InvestmentForex.trade_date <= end,
        )
    ).all()
    by_day: dict[int, dict] = {
        d: {"day": d, "pnl": 0.0, "count": 0, "win": 0, "loss": 0, "position": False}
        for d in range(1, end.day + 1)
    }
    month_pnl = 0.0
    trading_days = 0
    for t in trades:
        b = by_day.setdefault(t.trade_date.day, {"day": t.trade_date.day, "pnl": 0.0, "count": 0, "win": 0, "loss": 0, "position": False, })
        b["count"] += 1
        net = _trade_net(t)
        b["pnl"] = round(b["pnl"] + net, 2)
        month_pnl += net
        if net > 0:
            b["win"] += 1
        elif net < 0:
            b["loss"] += 1
        if t.status == "open":
            b["position"] = True

    # 月汇总：月盈亏、交易日数、盈利/亏损天数、月收益率
    month_pnl = round(month_pnl, 2)
    trading_days = sum(1 for d in by_day.values() if d["count"] > 0)
    win_days = sum(1 for d in by_day.values() if d["pnl"] > 0)
    loss_days = sum(1 for d in by_day.values() if d["pnl"] < 0)
    # 月收益率 = 本月净盈亏 / 累计净投入本金（入金+赠金-亏损-出金），该值也可由其余明细调整
    fund_rows = db.scalars(
        select(InvestmentFundRecord).where(InvestmentFundRecord.user_id == current_user.id)
    ).all()
    net_capital = sum(f.amount for f in fund_rows if f.record_type == "deposit")
    net_capital += sum(max(0, f.amount) for f in fund_rows if f.record_type == "experience")
    net_capital -= sum(f.amount for f in fund_rows if f.record_type == "withdraw")
    net_capital -= sum(-f.amount for f in fund_rows if f.record_type == "experience" and f.amount < 0 and not (f.note and "bns807" in f.note))
    base = net_capital if net_capital and net_capital != 0 else 1.0
    return_rate = round(month_pnl / base * 100, 2) if base else 0.0

    return {
        "year": y,
        "month": m,
        "days": sorted(by_day.values(), key=lambda x: x["day"]),
        "summary": {
            "month_pnl": month_pnl,
            "trading_days": trading_days,
            "win_days": win_days,
            "loss_days": loss_days,
            "return_rate": return_rate,
        },
    }


@router.get("/{item_id}", response_model=ForexRead)
def get_item(item_id: int, db: Session = Depends(get_db),
             current_user: UserProfile = Depends(get_current_user)):
    """置于 /stats /calendar 之后定义，避免 /stats 等字面路径被当作 item_id 捕获。"""
    obj = db.scalar(
        select(InvestmentForex).where(
            InvestmentForex.id == item_id,
            InvestmentForex.user_id == current_user.id,
        )
    )
    if not obj:
        raise HTTPException(status_code=404, detail="记录不存在")
    return _read(obj)


# --------------------------------------------------------------------------
# xlsx 导入（MT5 清洗后导出）
# --------------------------------------------------------------------------
_HEADER_MAP = {
    "日期时间": "trade_date", "交易品种": "symbol", "订单类型": "order_type",
    "开仓价格": "open_price", "手数": "lot_size", "手续费": "commission",
    "平仓价格": "close_price", "盈亏金额": "pnl", "隔夜费": "overnight_fee",
    "开仓时间": "open_time", "平仓时间": "close_time", "持仓时间": "holding",
    "备注": "note",
}


def _trade_unique_key(rec: InvestmentForex):
    """交易唯一键：优先按 MT5 订单号（备注中的 ID:xxx）判定，其次回退到业务字段。"""
    if rec.note:
        # 提取备注中的 ID:数字 —— 订单号是去重依据（不同订单可能业务字段相同）
        import re as _re
        m = _re.search(r"ID[:：]\s*([0-9]+)", rec.note)
        if m:
            return ("id", m.group(1))
    close_price = rec.close_price if rec.close_price is not None else ""
    pnl = rec.pnl if rec.pnl is not None else ""
    return (
        "biz",
        rec.trade_date,
        rec.symbol,
        rec.order_type,
        rec.open_price,
        rec.lot_size,
        close_price,
        pnl,
        rec.commission,
        rec.overnight_fee,
    )


def parse_trade_rows(rows, seen_keys: set | None = None) -> tuple[list[InvestmentForex], int]:
    """把 MT5 清洗后的原始行解析为交易记录列表。返回 (records, skipped)。

    表头：日期时间|交易品种|订单类型|开仓价格|手数|手续费|平仓价格|盈亏金额|隔夜费|开仓时间|平仓时间|持仓时间|备注
    （可含 ID 首列，自动忽略）

    支持去重：传入 seen_keys (set of unique keys) 时会跳过其中已存在的记录并计入 skipped；
    若不传，则只做文件内去重（同一文件内重复行）。
    """
    if not rows:
        return [], 0
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h in _HEADER_MAP:
            col_idx[_HEADER_MAP[h]] = i

    def get(raw, key):
        i = col_idx.get(key)
        if i is None or i >= len(raw):
            return None
        return raw[i]

    seen = set() if seen_keys is None else set(seen_keys)
    records, skipped = [], 0
    for raw in rows[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        symbol = get(raw, "symbol")
        day = _coerce_date(get(raw, "trade_date"))
        open_dt, close_dt = _combine_open_close(day, get(raw, "open_time"), get(raw, "close_time"))
        pnl = _num(get(raw, "pnl"))
        if not symbol or (open_dt is None and not get(raw, "trade_date")):
            skipped += 1
            continue

        order_type = str(get(raw, "order_type") or "").strip().lower()
        order_type = "buy" if order_type and order_type.startswith("buy") else ("sell" if order_type else "buy")

        close_price = _num(get(raw, "close_price"))
        # MT5 导入的数据均为已平仓
        status = "closed"

        holding = _holding_minutes(open_dt, close_dt) or _parse_duration_minutes(get(raw, "holding"))
        rec = InvestmentForex(
            trade_date=day or (open_dt.date() if open_dt else date.today()),
            symbol=str(symbol).strip(),
            order_type=order_type,
            open_price=_num(get(raw, "open_price"), 0) or 0,
            lot_size=_num(get(raw, "lot_size"), 0) or 0,
            commission=_num(get(raw, "commission"), 0) or 0,
            close_price=close_price,
            pnl=pnl,
            overnight_fee=_num(get(raw, "overnight_fee"), 0) or 0,
            open_time=open_dt,
            close_time=close_dt,
            holding=holding,
            status=status,
            note=str(get(raw, "note")).strip() if get(raw, "note") else None,
        )
        key = _trade_unique_key(rec)
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        records.append(rec)
    return records, skipped


@router.post("/import")
async def import_xlsx(
    mode: str = Query("append", pattern="^(append|replace)$"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserProfile = Depends(get_current_user),
):
    """导入 MT5 导出的 xlsx。表头：日期时间|交易品种|订单类型|开仓价格|手数|手续费|平仓价格|盈亏金额|隔夜费|开仓时间|平仓时间|持仓时间|备注"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端缺少 openpyxl 依赖")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件：{exc}")

    rows = list(ws.iter_rows(values_only=True))
    # 已入库记录的唯一键集合：用于跨文件去重（防止重复导入）
    existing_recs = db.scalars(
        select(InvestmentForex).where(InvestmentForex.user_id == current_user.id)
    ).all()
    existing_keys = {_trade_unique_key(r) for r in existing_recs}
    records, skipped = parse_trade_rows(rows, seen_keys=existing_keys)
    for rec in records:
        rec.user_id = current_user.id

    if records:
        if mode == "replace":
            db.query(InvestmentForex).filter(
                InvestmentForex.user_id == current_user.id
            ).delete()
        db.add_all(records)
        db.commit()
    return {"imported": len(records), "skipped": skipped}