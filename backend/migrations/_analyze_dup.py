# -*- coding: utf-8 -*-
"""分析数据导入_备份.xlsx 的重复模式"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook

XLSX = r"E:\Code\LifeOS V2.0\.trae\数据导入_备份.xlsx"
wb = load_workbook(XLSX, data_only=True, read_only=True)
print("sheets:", wb.sheetnames)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
header = [str(c).strip() if c is not None else "" for c in rows[0]]
print("表头:", header)
print("总行数(含表头):", len(rows))
data = rows[1:]
print("数据行:", len(data))

# 候选去重键：整行 / 备注 / 开平仓+品种
from collections import Counter
def key_row(r):
    return tuple(str(x) for x in r)

# 按整行去重
c_row = Counter(key_row(r) for r in data)
print("\n按整行去重: 唯一", len(c_row), "重复行数(去重前)", sum(v-1 for v in c_row.values() if v>1))

# 常用去重键：备注 或 开仓时间+品种+开仓价+手数
import re
def trade_remark(r):
    # 假设倒数第几列是备注。找"备注"列
    return None

# 找列索引
hmap = {}
for i,h in enumerate(header):
    hmap[h] = i
print("列索引:", hmap)

# 候选：开仓时间(纯时间) + 品种 + 手数 + 开仓价 作为业务唯一键
def biz_key(r):
    parts = []
    for col in ["开仓时间","交易品种","手数","开仓价格","订单类型","日期时间"]:
        if col in hmap:
            v = r[hmap[col]]
            parts.append(str(v).strip() if v is not None else "")
    return tuple(parts)

c_biz = Counter(biz_key(r) for r in data)
dup_biz = {k:v for k,v in c_biz.items() if v>1}
print("\n按[开仓时间+品种+手数+开仓价+方向+日期]去重: 唯一", len(c_biz), "重复组数", len(dup_biz), "重复条数", sum(v-1 for v in dup_biz.values()))
# 展示几组重复
for i,(k,v) in enumerate(dup_biz.items()):
    if i>=3: break
    print("  重复组:", k, "出现", v, "次")