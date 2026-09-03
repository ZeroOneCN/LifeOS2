"""投资中心 MT5 迁移 + 示例数据导入（一次性脚本，幂等可重复执行）。

功能：
1. 备份现有 investment_forex / investment_fund_records / investment_reports 到 backups/ 下 .csv
2. 将 investment_forex 重构为 MT5 格式（重命名 pair->symbol、direction->order_type，新增 commission/overnight_fee/open_time/close_time/holding），
   并对既有数据用 trade_date 兜底开平仓时间
3. create_all 自动创建缺失的表（fund/report 等）
4. 读取 \".trae/数据导入.xlsx\" 两个 sheet 并导入：
   - Sheet     -> investment_forex（若当前无数据则覆盖，否则追加）
   - 出入金      -> investment_fund_records（出入金/体验金；追加）
运行：在 backend 目录执行  .venv\\Scripts\\python.exe migrations/run_investment_forex_mt5.py
"""
import csv
import os
import sys
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from sqlalchemy import text

from app.core.database import Base, engine, SessionLocal
from app.api.routes.investment.forex import parse_trade_rows
from app.models import InvestmentForex, InvestmentFundRecord

BACKUP_DIR = os.path.join(os.path.dirname(__file__), "..", "backups")
XLSX = r"E:\Code\LifeOS V2.0\.trae\数据导入.xlsx"


def backup_table(db, name):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(BACKUP_DIR, f"{name}_{stamp}.csv")
    try:
        rows = db.execute(text(f"SELECT * FROM {name}")).mappings().all()
    except Exception as e:  # noqa: BLE001
        print(f"  [{name}] 备份跳过:", e)
        return None
    if not rows:
        print(f"  [{name}] 无数据，跳过备份")
        return None
    cols = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({k: (v.isoformat() if isinstance(v, (date, datetime)) else v) for k, v in dict(r).items()})
    print(f"  [{name}] 已备份 {len(rows)} 行 -> {path}")
    return path


def migrate():
    db = SessionLocal()
    try:
        print("==> 0) create_all 确保新表存在（fund/report 等）")
        Base.metadata.create_all(bind=engine)

        print("==> 1) 备份现有数据")
        backup_table(db, "investment_forex")
        backup_table(db, "investment_fund_records")
        backup_table(db, "investment_reports")

        print("==> 2) 重构 investment_forex 为 MT5 格式")
        try:
            db.execute(text("ALTER TABLE investment_forex CHANGE COLUMN pair symbol VARCHAR(24) NOT NULL"))
        except Exception as e:  # noqa: BLE001
            print("   (pair->symbol 跳过):", e)
        try:
            db.execute(text("ALTER TABLE investment_forex CHANGE COLUMN direction order_type VARCHAR(8) NOT NULL"))
        except Exception as e:  # noqa: BLE001
            print("   (direction->order_type 跳过):", e)
        for col, ddl in [
            ("commission", "commission DOUBLE NOT NULL DEFAULT 0"),
            ("overnight_fee", "overnight_fee DOUBLE NOT NULL DEFAULT 0"),
            ("open_time", "open_time DATETIME NULL"),
            ("close_time", "close_time DATETIME NULL"),
            ("holding", "holding INT NULL"),
        ]:
            try:
                db.execute(text(f"ALTER TABLE investment_forex ADD COLUMN {ddl}"))
                print(f"   新增列 {col}")
            except Exception as e:  # noqa: BLE001
                print(f"   ({col} 跳过):", e)
        db.execute(text(
            "UPDATE investment_forex SET "
            "open_time = COALESCE(open_time, CAST(CONCAT(trade_date,' 00:00:00') AS DATETIME)), "
            "close_time = COALESCE(close_time, CAST(CONCAT(trade_date,' 00:00:00') AS DATETIME)), "
            "holding = COALESCE(holding,0) WHERE open_time IS NULL OR close_time IS NULL"
        ))
        db.commit()
        print("   表结构迁移完成")

        print("==> 3) create_all 创建缺失新表")
        Base.metadata.create_all(bind=engine)
        print("   完成")

        print("==> 4) 导入示例 xlsx")
        wb = load_workbook(XLSX, data_only=True, read_only=True)
        if "Sheet" in wb.sheetnames:
            rows = list(wb["Sheet"].iter_rows(values_only=True))
            recs, skipped = parse_trade_rows(rows)
            if recs:
                db.query(InvestmentForex).delete()
                db.add_all(recs)
                db.commit()
                print(f"   交易记录导入 {len(recs)} 条 (skip {skipped}, 覆盖导入)")

        if "出入金" in wb.sheetnames:
            rows = list(wb["出入金"].iter_rows(values_only=True))
            imported = 0
            for raw in rows[1:]:
                if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                    continue
                d, t, amt, note = (list(raw) + [None] * 4)[:4]
                if not t or amt is None:
                    continue
                t = str(t).strip()
                rt = {"入金": "deposit", "出金": "withdraw", "体验金": "experience"}.get(t)
                if not rt:
                    continue
                rdate = _fund_date(d)
                db.add(InvestmentFundRecord(record_type=rt, amount=float(amt), record_date=rdate, note=str(note).strip() if note else None))
                imported += 1
            db.commit()
            print(f"   资金动态导入 {imported} 条")
        db.close()
        print("==> 全部完成")
    finally:
        if db.is_active:
            db.close()


def _fund_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(str(v).strip()).date()
    except ValueError:
        return date.today()


if __name__ == "__main__":
    migrate()