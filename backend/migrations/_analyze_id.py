# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from collections import Counter
XLSX = r"E:\Code\LifeOS V2.0\.trae\数据导入_备份.xlsx"
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
data = rows[1:]
header = [str(c).strip() if c is not None else "" for c in rows[0]]
hid = 0  # ID col

c_id = Counter(str(r[hid]).strip() for r in data)
dup_id = {k: v for k, v in c_id.items() if v > 1}
print("按 ID 去重: 唯一", len(c_id), "重复组", len(dup_id), "重复条数", sum(v-1 for v in dup_id.values()))
for k, v in list(dup_id.items())[:5]:
    print("  重复ID", k, "次", v)

# 检查ID为空的情况
empty_id = sum(1 for r in data if str(r[hid]).strip()=="")
print("空ID行:", empty_id)

# 检查整数行去重和ID去重差异
c_row = Counter(tuple(str(x) for x in r) for r in data)
print("\n整行去重：唯一", len(c_row), "重复", sum(v-1 for v in c_row.values() if v>1))

# 是否ID唯一即可覆盖大部分
# 显示：ID相同但其他字段不同的情况
id2rows = {}
from collections import defaultdict
grp = defaultdict(list)
for r in data:
    grp[str(r[hid]).strip()].append(tuple(str(x) for x in r))
diff_under_same_id = 0
for k, v in grp.items():
    if len(set(v)) > 1 and len(v) > 1:
        diff_under_same_id += 1
print("\n相同ID但整行不完全相同的组数:", diff_under_same_id)